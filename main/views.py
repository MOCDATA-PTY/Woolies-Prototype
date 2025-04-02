# Django Core Imports
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm, SetPasswordForm
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models import Sum, Avg, Count, Min, Max, StdDev, F, Q
from django.db.models.functions import TruncWeek, TruncMonth
from prophet import Prophet

# Local Forms & Models
from .forms import EggFarmForm, UserForm, CustomUserCreationForm
from .models import EggFarm

# Forecasting & Analysis
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
import logging

from sklearn.linear_model import LinearRegression, Ridge
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error, mean_absolute_percentage_error
from sklearn.model_selection import cross_val_score


# Set up logger
logger = logging.getLogger(__name__)


# File Handling
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from io import StringIO, BytesIO, TextIOWrapper
from zipfile import ZipFile

# Scientific & Statistical Computing
from scipy import stats

# Utilities
import traceback
import logging

# Custom Decorators
from .decorators import critical_view_decorator

# Logging Configuration
logger = logging.getLogger(__name__)


from prophet import Prophet
from django.http import JsonResponse
from .models import EggFarm
from django.db.models import Sum
from django.db.models.functions import TruncWeek
import pandas as pd
from datetime import timedelta

def prophet_forecast_view(request):
    # Step 1: Get historical data grouped by week
    data = (
        EggFarm.objects
        .filter(production_date__isnull=False)
        .annotate(week=TruncWeek('production_date'))
        .values('week')
        .annotate(total=Sum('weekly_egg_production'))
        .order_by('week')
    )

    if not data:
        return JsonResponse({'error': 'No data available for forecasting'}, status=400)

    # Step 2: Convert to DataFrame for Prophet
    df = pd.DataFrame(data)
    df.rename(columns={'week': 'ds', 'total': 'y'}, inplace=True)
    df['ds'] = pd.to_datetime(df['ds'])
    df = df.sort_values('ds')

    # Step 3: Initialize and train Prophet model
    model = Prophet(weekly_seasonality=True, yearly_seasonality=False, daily_seasonality=False)
    model.fit(df)

    # Step 4: Create future dataframe
    future = model.make_future_dataframe(periods=12, freq='W')  # Forecast 12 weeks into the future
    forecast = model.predict(future)

    # Step 5: Format data for Chart.js
    actual_data = df['y'].round(2).tolist()
    forecast_data = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(12)

    result = {
        'labels': forecast['ds'].dt.strftime('%Y-%m-%d').tolist(),
        'actual': [None] * (len(forecast) - len(df)) + actual_data,
        'forecast': forecast['yhat'].round(2).tolist(),
        'lower_bound': forecast['yhat_lower'].round(2).tolist(),
        'upper_bound': forecast['yhat_upper'].round(2).tolist()
    }

    return JsonResponse(result)

# Helper Functions
def clear_messages(request):
    """
    Helper function to clear all messages from the request.
    """
    storage = messages.get_messages(request)
    for message in storage:
        pass
    storage.used = True

def parse_date(date_val):
    """
    Safely parse Excel date values into Python date objects.
    Accepts strings or datetime objects.
    """
    from datetime import datetime, date

    if not date_val:
        return None

    if isinstance(date_val, (datetime, date)):
        return date_val.date() if isinstance(date_val, datetime) else date_val

    try:
        return datetime.strptime(str(date_val), '%Y-%m-%d').date()
    except ValueError:
        return None

# Authentication Views
@csrf_protect
def register_user(request):
    """
    Handle user registration with enhanced security.
    """
    # Force logout of any existing session
    if request.user.is_authenticated:
        logout(request)
    
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Create user with additional security steps
            user = form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}. You can now log in.')
            return redirect('login')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'accounts/register.html', {'form': form})

@never_cache
@csrf_protect
def login_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')  # Go to home/dashboard
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'accounts/login.html')

def logout_user(request):
    """
    Handle user logout with enhanced security.
    """
    # Clear all messages
    storage = messages.get_messages(request)
    storage.used = True
    
    # Completely invalidate the session
    request.session.flush()
    
    # Logout and redirect
    logout(request)
    
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')

# User Management Views
@user_passes_test(lambda u: u.is_staff, login_url='login')
def create_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = form.cleaned_data['is_staff']
            user.save()
            messages.success(request, 'User created successfully.')
            return redirect('user_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'accounts/create_user.html', {'form': form})

@staff_member_required(login_url='login')
def user_list(request):
    users = User.objects.all()
    return render(request, 'accounts/user_list.html', {'users': users})

@user_passes_test(lambda u: u.is_staff, login_url='login')
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        user.is_staff = bool(request.POST.get('is_staff'))
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = UserForm(instance=user)
    
    return render(request, 'accounts/edit_user.html', {'form': form, 'user': user})

def edit_user_password(request, user_id):
    user = get_object_or_404(User, id=user_id)
    form = SetPasswordForm(user, request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Password updated successfully.')
            return redirect('user_list')
        else:
            messages.error(request, 'Please correct the errors below.')

    return render(request, 'accounts/edit_password.html', {'form': form, 'user_obj': user})

@login_required(login_url='login')
def delete_user(request, user_id):
    """
    View to handle user deletion with confirmation.
    GET request shows confirmation page, POST request performs deletion.
    """
    # Get the user object or return 404 if not found
    user = get_object_or_404(User, id=user_id)

    # Prevent self-deletion (important security check)
    if request.user == user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('user_list')

    # Handle POST request (user confirmed deletion)
    if request.method == 'POST':
        # Store username for message
        username = user.username
        
        # Delete the user
        user.delete()
        
        # Add success message
        messages.success(request, f'User "{username}" has been successfully deleted.')
        
        # Redirect to user list
        return redirect('user_list')
    
    # Handle GET request (show confirmation page)
    return render(request, 'accounts/confirm_delete.html', {'user_obj': user})
# Main Views
@login_required(login_url='login')
def home(request):
    print("User authenticated:", request.user.is_authenticated)
    return render(request, 'main/home.html')

@login_required
def test_dashboard(request):
    return render(request, 'main/dashboard_test.html')

# Egg Farm CRUD Views
@critical_view_decorator
@login_required(login_url='login')
def add_egg_farm(request):
    """
    View to add a new egg farm record.
    Displays existing farm names for reference.
    """
    existing_farms = EggFarm.objects.values_list('farm_name', flat=True).distinct()

    if request.method == 'POST':
        form = EggFarmForm(request.POST)
        if form.is_valid():
            # Ensure we're not duplicating a record with the same farm/site/house/date
            farm_name = form.cleaned_data['farm_name']
            site_name = form.cleaned_data['site_name']
            hen_house = form.cleaned_data['hen_house_number']
            prod_date = form.cleaned_data['production_date']
            
            existing = EggFarm.objects.filter(
                farm_name=farm_name,
                site_name=site_name,
                hen_house_number=hen_house,
                production_date=prod_date
            ).exists()
            
            if existing:
                messages.error(request, 'A record already exists for this farm location and date.')
            else:
                form.save()
                messages.success(request, 'Egg farm record added successfully.')
                return redirect('egg_farm_list')
        else:
            print("Form errors:", form.errors)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EggFarmForm()

    return render(request, 'main/add_egg_farm.html', {
        'form': form,
        'farms': existing_farms
    })

@critical_view_decorator
@login_required(login_url='login')
def egg_farm_list(request):
    """
    List all egg farm records with comprehensive filtering options.
    """
    clear_messages(request)
    
    # Fetch all farms, not just filtered
    all_farms = EggFarm.objects.all()

    # Get all possible filter options
    all_supplier_names = sorted(set(all_farms.values_list('supplier_name', flat=True)))
    all_farm_names = sorted(set(all_farms.values_list('farm_name', flat=True)))
    all_site_names = sorted(set(all_farms.values_list('site_name', flat=True)))
    
    # Default queryset for displaying farms
    egg_farms = EggFarm.objects.all()
    
    # Advanced filtering
    supplier_name = request.GET.get('supplier_name')
    farm_name = request.GET.get('farm_name')
    site_name = request.GET.get('site_name')
    breed = request.GET.get('breed')
    health_issues = request.GET.get('health_issues')
    placement_from = request.GET.get('placement_from')
    placement_to = request.GET.get('placement_to')
    min_hens = request.GET.get('min_hens')
    max_hens = request.GET.get('max_hens')
    min_age = request.GET.get('min_age')
    max_age = request.GET.get('max_age')

    # Apply filters if provided
    if supplier_name:
        egg_farms = egg_farms.filter(supplier_name__icontains=supplier_name)
    if farm_name:
        egg_farms = egg_farms.filter(farm_name__icontains=farm_name)
    if site_name:
        egg_farms = egg_farms.filter(site_name__icontains=site_name)
    if breed:
        egg_farms = egg_farms.filter(breed=breed)
    if health_issues:
        egg_farms = egg_farms.filter(health_issues=health_issues)
    if placement_from:
        egg_farms = egg_farms.filter(recent_placement_date__gte=placement_from)
    if placement_to:
        egg_farms = egg_farms.filter(recent_placement_date__lte=placement_to)
    if min_hens:
        egg_farms = egg_farms.filter(current_number_of_hens__gte=int(min_hens))
    if max_hens:
        egg_farms = egg_farms.filter(current_number_of_hens__lte=int(max_hens))
    if min_age:
        egg_farms = egg_farms.filter(average_age_weeks__gte=float(min_age))
    if max_age:
        egg_farms = egg_farms.filter(average_age_weeks__lte=float(max_age))
    
    # Default ordering
    egg_farms = egg_farms.order_by('-production_date', 'farm_name', 'site_name', 'hen_house_number')

    context = {
        'egg_farms': egg_farms,
        
        # Filter options
        'supplier_names': all_supplier_names,
        'farm_names': all_farm_names,
        'site_names': all_site_names,
        
        # Breed and health options
        'breed_choices': EggFarm.BREED_CHOICES,
        'health_choices': EggFarm.HEALTH_CHOICES,
        
        # Current filter values to maintain form state
        'current_filters': {
            'supplier_name': supplier_name,
            'farm_name': farm_name,
            'site_name': site_name,
            'breed': breed,
            'health_issues': health_issues,
            'placement_from': placement_from,
            'placement_to': placement_to,
            'min_hens': min_hens,
            'max_hens': max_hens,
            'min_age': min_age,
            'max_age': max_age,
        }
    }

    return render(request, 'main/egg_farm_list.html', context)
@critical_view_decorator
@login_required(login_url='login')
def edit_egg_farm(request, pk):
    """
    Edit an existing egg farm record.
    Allows keeping the original record or updating.
    """
    clear_messages(request)
    egg_farm = get_object_or_404(EggFarm, pk=pk)
    
    # Option to keep original record
    if request.method == 'POST' and 'keep_original' in request.POST:
        messages.info(request, 'No changes were made to the record.')
        return redirect('egg_farm_list')
    
    if request.method == 'POST':
        form = EggFarmForm(request.POST, instance=egg_farm)
        if form.is_valid():
            # Check if we're creating a duplicate after edit
            farm_name = form.cleaned_data['farm_name']
            site_name = form.cleaned_data['site_name']
            hen_house = form.cleaned_data['hen_house_number']
            prod_date = form.cleaned_data['production_date']
            
            existing = EggFarm.objects.filter(
                farm_name=farm_name,
                site_name=site_name,
                hen_house_number=hen_house,
                production_date=prod_date
            ).exclude(pk=pk).exists()
            
            if existing:
                messages.error(request, 'A record already exists for this farm location and date.')
            else:
                form.save()
                messages.success(request, 'Egg farm record updated successfully.')
                return redirect('egg_farm_list')
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        form = EggFarmForm(instance=egg_farm)
    return render(request, 'main/edit_egg_farm.html', {'form': form, 'egg_farm': egg_farm})

@critical_view_decorator
@login_required(login_url='login')
def delete_egg_farm(request, pk):
    """
    Delete a specific egg farm record.
    """
    clear_messages(request)
    egg_farm = get_object_or_404(EggFarm, pk=pk)
    if request.method == 'POST':
        egg_farm.delete()
        messages.success(request, 'Egg farm record deleted successfully.')
    return redirect('egg_farm_list')

# Import/Export Views
@critical_view_decorator
@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def import_egg_farms(request):
    if request.method == 'POST' and request.FILES.get('data_file'):
        data_file = request.FILES['data_file']
        filename = data_file.name.lower()

        try:
            rows = []

            # ✅ Excel (.xlsx)
            if filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(data_file)
                sheet = wb.active
                rows = list(sheet.iter_rows(min_row=2, values_only=True))

            # ✅ CSV (.csv)
            elif filename.endswith('.csv'):
                decoded_file = data_file.read().decode('utf-8')
                lines = decoded_file.splitlines()

                if not lines:
                    raise ValueError("CSV file is empty.")

                reader = csv.reader(lines)
                header = next(reader, None)  # safely try to get header

                if not header:
                    raise ValueError("CSV file has no header row.")

                rows = list(reader)
                if not rows:
                    raise ValueError("CSV file contains only a header and no data.")

            else:
                messages.error(request, 'Unsupported file format. Please upload a .xlsx or .csv file.')
                return redirect('egg_farm_list')

            # ✅ Process rows
            results = process_import_rows(rows)
            generate_import_messages(request, results)
            return redirect('egg_farm_list')

        except Exception as e:
            error_message = f"Error processing file: {str(e)}"
            print("❌ FILE IMPORT ERROR ❌")
            traceback.print_exc()  # Print full stack trace
            messages.error(request, error_message)
            return redirect('egg_farm_list')

    return render(request, 'main/import_egg_farms.html')

def process_import_rows(rows):
    """
    Process imported rows of data, creating new records for each row.
    Now respects the unique constraint on farm/site/house/date to avoid duplicates.
    """
    results = {
        'created': [],
        'updated': [],
        'skipped': [],
        'errors': []
    }

    for index, row in enumerate(rows):
        try:
            cleaned_data = validate_egg_farm_data(row)
            if not cleaned_data:
                results['errors'].append(f"[Row {index + 2}] Validation failed: {row}")
                continue

            with transaction.atomic():  # 🔐 Each row safely in its own transaction
                # Check if a record with this combination exists
                existing = EggFarm.objects.filter(
                    supplier_name=cleaned_data['supplier_name'],
                    farm_name=cleaned_data['farm_name'],
                    site_name=cleaned_data['site_name'],
                    hen_house_number=cleaned_data['hen_house_number'],
                    production_date=cleaned_data['production_date']
                ).first()

                if existing:
                    # Update the existing record
                    for key, value in cleaned_data.items():
                        setattr(existing, key, value)
                    existing.save()
                    results['updated'].append(existing)
                else:
                    # Create a new record
                    new_farm = EggFarm.objects.create(**cleaned_data)
                    results['created'].append(new_farm)

        except Exception as e:
            results['errors'].append(f"[Row {index + 2}] Processing error: {e} | Data: {row}")

    return results

def generate_import_messages(request, results, debug=True):
    """
    Generate user-friendly messages based on import results.
    Prints error entries to console if debug=True.
    """
    if results['created']:
        messages.success(request, f'{len(results["created"])} new entries created.')
    
    if results['updated']:
        messages.info(request, f'{len(results["updated"])} entries updated.')
    
    if results['skipped']:
        messages.warning(request, f'{len(results["skipped"])} entries skipped.')

    if results['errors']:
        messages.error(
            request,
            f'{len(results["errors"])} entries could not be processed. '
            'Please check your file format and data.'
        )

        if debug:
            print("❌ DEBUG: Error Entries")
            for i, err in enumerate(results['errors'], 1):
                print(f"Error {i}: {err}")

    if not any(results.values()):
        messages.warning(request, 'No entries were imported. File might be empty.')

@critical_view_decorator
@login_required(login_url='login')
def export_egg_farms(request):
    # In-memory ZIP file
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        data = EggFarm.objects.all()

        # ====== 1. Create CSV content ======
        csv_string_io = StringIO()
        csv_writer = csv.writer(csv_string_io)

        header = [
            'Supplier Name', 'Farm Name', 'Site Name', 'Hen House Number',
            'Current Number of Hens', 'Average Age (Weeks)', 'Breed', 'Breed Percentage',
            'Daily Egg Production', 'Weekly Egg Production',
            'Small Eggs %', 'Medium Eggs %', 'Large Eggs %',
            'XL Eggs %', 'Jumbo Eggs %', 'Defective Eggs %', 'First Grade Egg %',
            'Avg Feed Consumption', 'Avg Water Intake',
            'Weekly Mortality Rate', 'Total Mortality', 'Hens Culled',
            'Expected Cull Date', 'Recent Placement Date', 'Age at Placement',
            'Upcoming Placement Date', 'Health Issues', 'Recent Vaccinations',
            'Next Vaccination Date', 'Time to Distribution', 'Market Condition Notes',
            'Weekly Production Trend', 'Seasonal Trend Notes', 'Production Date'
        ]
        csv_writer.writerow(header)

        for farm in data:
            csv_writer.writerow([
                farm.supplier_name, farm.farm_name, farm.site_name, farm.hen_house_number,
                farm.current_number_of_hens, farm.average_age_weeks, farm.breed, farm.breed_percentage,
                farm.daily_egg_production, farm.weekly_egg_production,
                farm.small_eggs_percentage, farm.medium_eggs_percentage, farm.large_eggs_percentage,
                farm.extra_large_eggs_percentage, farm.jumbo_eggs_percentage, farm.defective_eggs_percentage,
                farm.first_grade_egg_percentage, farm.average_feed_consumption, farm.average_water_intake,
                farm.weekly_mortality_rate, farm.total_mortality, farm.hens_culled,
                farm.expected_cull_date, farm.recent_placement_date, farm.age_of_pullets_at_placement,
                farm.upcoming_placement_date, farm.health_issues, farm.recent_vaccinations,
                farm.next_vaccination_date, farm.time_to_distribution, farm.market_condition_notes,
                farm.weekly_production_trend, farm.seasonal_trend_notes, farm.production_date
            ])

        # Add CSV to ZIP
        zip_file.writestr("egg_farms.csv", csv_string_io.getvalue())

        # ====== 2. Create Excel (.xlsx) content ======
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Egg Farms"
        ws.append(header)

        for farm in data:
            ws.append([
                farm.supplier_name, farm.farm_name, farm.site_name, farm.hen_house_number,
                farm.current_number_of_hens, farm.average_age_weeks, farm.breed, farm.breed_percentage,
                farm.daily_egg_production, farm.weekly_egg_production,
                farm.small_eggs_percentage, farm.medium_eggs_percentage, farm.large_eggs_percentage,
                farm.extra_large_eggs_percentage, farm.jumbo_eggs_percentage, farm.defective_eggs_percentage,
                farm.first_grade_egg_percentage, farm.average_feed_consumption, farm.average_water_intake,
                farm.weekly_mortality_rate, farm.total_mortality, farm.hens_culled,
                farm.expected_cull_date, farm.recent_placement_date, farm.age_of_pullets_at_placement,
                farm.upcoming_placement_date, farm.health_issues, farm.recent_vaccinations,
                farm.next_vaccination_date, farm.time_to_distribution, farm.market_condition_notes,
                farm.weekly_production_trend, farm.seasonal_trend_notes, farm.production_date
            ])

        # Auto-fit Excel columns
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zip_file.writestr("egg_farms.xlsx", excel_buffer.read())

    # Return ZIP Response
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="egg_farms_export.zip"'
    return response

@critical_view_decorator
@login_required(login_url='login')
def clear_database(request):
    """
    Clear all egg farm records from the database.
    Requires user authentication.
    """
    if request.method == 'POST':
        with transaction.atomic():
            EggFarm.objects.all().delete()
            messages.success(request, "All egg farm records have been successfully deleted.")
        return redirect('egg_farm_list')
    else:
        messages.error(request, "Invalid request method.")
        return redirect('egg_farm_list')

def validate_egg_farm_data(row):
    """
    Validate and clean imported data row.
    Returns a cleaned dictionary or None if invalid.
    """
    try:
        return {
            'supplier_name': str(row[0]).strip(),
            'farm_name': str(row[1]).strip(),
            'site_name': str(row[2]).strip(),
            'hen_house_number': str(row[3]).strip(),
            'current_number_of_hens': int(row[4]),
            'average_age_weeks': float(row[5]),
            'breed': str(row[6]).strip(),
            'breed_percentage': float(row[7]),
            'daily_egg_production': float(row[8]),
            'weekly_egg_production': float(row[9]),
            'small_eggs_percentage': float(row[10]),
            'medium_eggs_percentage': float(row[11]),
            'large_eggs_percentage': float(row[12]),
            'extra_large_eggs_percentage': float(row[13]),
            'jumbo_eggs_percentage': float(row[14]),
            'defective_eggs_percentage': float(row[15]),
            'first_grade_egg_percentage': float(row[16]),
            'average_feed_consumption': float(row[17]),
            'average_water_intake': float(row[18]),
            'weekly_mortality_rate': float(row[19]),
            'total_mortality': int(row[20]),
            'hens_culled': int(row[21]),
            'expected_cull_date': parse_date(row[22]),
            'recent_placement_date': parse_date(row[23]),
            'age_of_pullets_at_placement': float(row[24]),
            'upcoming_placement_date': parse_date(row[25]),
            'health_issues': str(row[26]).strip(),
            'recent_vaccinations': str(row[27]).strip(),
            'next_vaccination_date': parse_date(row[28]),
            'time_to_distribution': int(row[29]),
            'market_condition_notes': str(row[30]).strip(),
            'weekly_production_trend': float(row[31]),
            'seasonal_trend_notes': str(row[32]).strip(),
            'production_date': parse_date(row[33])  # ✅ Added production_date here
        }
    except Exception as e:
        print(f"⚠️ Row validation exception: {e}")
        return None

# Dashboard and Analytics Views
@critical_view_decorator
@login_required(login_url='login')
def dashboard(request):
    """
    Render the working dashboard page with summarized stats and filters.
    """
    egg_farms = EggFarm.objects.all()

    # Filters
    farm = request.GET.get('farm')
    site_name = request.GET.get('site_name')
    breed = request.GET.get('breed')
    min_hens = request.GET.get('min_hens')
    max_hens = request.GET.get('max_hens')
    health_status = request.GET.get('health_status')
    min_age = request.GET.get('min_age')
    max_age = request.GET.get('max_age')

    # Apply filters
    if farm:
        egg_farms = egg_farms.filter(farm_name__icontains=farm)
    if site_name:
        egg_farms = egg_farms.filter(site_name__icontains=site_name)
    if breed:
        egg_farms = egg_farms.filter(breed=breed)
    if min_hens:
        egg_farms = egg_farms.filter(current_number_of_hens__gte=int(min_hens))
    if max_hens:
        egg_farms = egg_farms.filter(current_number_of_hens__lte=int(max_hens))
    if health_status:
        egg_farms = egg_farms.filter(health_issues=health_status)
    if min_age:
        egg_farms = egg_farms.filter(average_age_weeks__gte=float(min_age))
    if max_age:
        egg_farms = egg_farms.filter(average_age_weeks__lte=float(max_age))

    # Get the most recent data for KPIs - using a subquery to get the max date for each farm
    latest_records = {}
    for farm_key in egg_farms.values('farm_name', 'site_name', 'hen_house_number').distinct():
        latest_date = egg_farms.filter(
            farm_name=farm_key['farm_name'],
            site_name=farm_key['site_name'],
            hen_house_number=farm_key['hen_house_number']
        ).order_by('-production_date').first()
        
        if latest_date:
            key = (farm_key['farm_name'], farm_key['site_name'], farm_key['hen_house_number'])
            latest_records[key] = latest_date.id

    # Use the latest records for KPI calculations
    latest_ids = list(latest_records.values())
    latest_farms = EggFarm.objects.filter(id__in=latest_ids)

    total_farms = latest_farms.values('farm_name').distinct().count()
    total_hens = latest_farms.aggregate(Sum('current_number_of_hens'))['current_number_of_hens__sum'] or 0
    average_age = latest_farms.aggregate(Avg('average_age_weeks'))['average_age_weeks__avg'] or 0
    total_weekly_production = latest_farms.aggregate(Sum('weekly_egg_production'))['weekly_egg_production__sum'] or 0

    egg_size_distribution = {
        'small': latest_farms.aggregate(Avg('small_eggs_percentage'))['small_eggs_percentage__avg'] or 0,
        'medium': latest_farms.aggregate(Avg('medium_eggs_percentage'))['medium_eggs_percentage__avg'] or 0,
        'large': latest_farms.aggregate(Avg('large_eggs_percentage'))['large_eggs_percentage__avg'] or 0,
        'extra_large': latest_farms.aggregate(Avg('extra_large_eggs_percentage'))['extra_large_eggs_percentage__avg'] or 0,
        'jumbo': latest_farms.aggregate(Avg('jumbo_eggs_percentage'))['jumbo_eggs_percentage__avg'] or 0,
    }

    context = {
        'total_farms': total_farms,
        'total_hens': total_hens,
        'average_age': round(average_age, 1),
        'total_weekly_production': total_weekly_production,
        'egg_size_distribution': egg_size_distribution,
        'farm_names': EggFarm.objects.values_list('farm_name', flat=True).distinct().order_by('farm_name'),
        'site_names': EggFarm.objects.values_list('site_name', flat=True).distinct().order_by('site_name'),
        'breed_choices': EggFarm.BREED_CHOICES,
        'health_choices': EggFarm.HEALTH_CHOICES,
    }

    return render(request, 'main/dashboard.html', context)

@critical_view_decorator
@login_required(login_url='login')
def dashboard_analytics_data(request):
    """
    Provide analytics data for dashboard.
    """
    # Get filters
    farm = request.GET.get('farm', None)
    site = request.GET.get('site_name', None)
    breed = request.GET.get('breed', None)
    health_status = request.GET.get('health_status', None)
    
    # Start with all farms
    queryset = EggFarm.objects.all()

    # Apply filters
    if farm:
        queryset = queryset.filter(farm_name__icontains=farm)
    if site:
        queryset = queryset.filter(site_name__icontains=site)
    if breed:
        queryset = queryset.filter(breed=breed)
    if health_status:
        queryset = queryset.filter(health_issues=health_status)
        
    # Get latest data for each farm/site/house combination
    latest_records = {}
    for farm_key in queryset.values('farm_name', 'site_name', 'hen_house_number').distinct():
        latest = queryset.filter(
            farm_name=farm_key['farm_name'],
            site_name=farm_key['site_name'],
            hen_house_number=farm_key['hen_house_number']
        ).order_by('-production_date').first()
        
        if latest:
            key = (farm_key['farm_name'], farm_key['site_name'], farm_key['hen_house_number'])
            latest_records[key] = latest.id
    
    latest_ids = list(latest_records.values())
    latest_data = EggFarm.objects.filter(id__in=latest_ids)

    # Feed vs Water
    feed_vs_water = list(latest_data.values_list('average_feed_consumption', 'average_water_intake'))

    # Breed Distribution
    breed_counts = latest_data.values('breed').annotate(total=Count('id')).order_by('-total')
    breed_distribution = {
        'labels': [b['breed'] for b in breed_counts],
        'data': [b['total'] for b in breed_counts]
    }

    # Mortality and Cull Rate per Breed
    mortality_data = latest_data.values('breed').annotate(
        avg_mortality=Avg('weekly_mortality_rate'),
        total_culled=Sum('hens_culled')
    )
    mortality_cull = {
        'labels': [m['breed'] for m in mortality_data],
        'mortality': [round(m['avg_mortality'] or 0, 2) for m in mortality_data],
        'culled': [m['total_culled'] or 0 for m in mortality_data]
    }

    # Egg Quality
    quality = latest_data.aggregate(
        defective=Avg('defective_eggs_percentage'),
        first_grade=Avg('first_grade_egg_percentage')
    )
    egg_quality = {
        'defective': round(quality['defective'] or 0, 2),
        'first_grade': round(quality['first_grade'] or 0, 2)
    }

    return JsonResponse({
        'feed_vs_water': feed_vs_water,
        'breed_distribution': breed_distribution,
        'mortality_cull': mortality_cull,
        'egg_quality': egg_quality
    })

@critical_view_decorator
@login_required(login_url='login')
def weekly_production_data(request):
    """
    Get weekly egg production data for charts
    """
    data = (
        EggFarm.objects
        .annotate(week=TruncWeek('production_date'))
        .values('week')
        .annotate(total_production=Sum('weekly_egg_production'))
        .order_by('week')
    )

    chart_data = {
        'labels': [d['week'].strftime("%Y-%m-%d") if d['week'] else "Unknown" for d in data],
        'data': [d['total_production'] for d in data]
    }
    return JsonResponse(chart_data)

@critical_view_decorator
@login_required(login_url='login')
def weekly_forecast_data(request):
    """
    Generate weekly egg production forecasts based on historical farm data.
    Supports multiple forecasting methods for comparison and longer time horizons.
    Enhanced with more robust statistical methods and outlier handling.
    """
    # Get filter parameters from request
    farm = request.GET.get('farm')
    site = request.GET.get('site_name')
    breed = request.GET.get('breed')
    health = request.GET.get('health_status')
    range_val = request.GET.get('range', '12w')  # Default to 12 weeks forecast
    method = request.GET.get('method', 'ensemble')  # Default to ensemble method

    # Convert range parameter to number of weeks
    try:
        if range_val.endswith('m'):
            # Convert months to weeks (more precisely)
            future_weeks = int(range_val.replace('m', '')) * 4.33  # Average weeks per month
            future_weeks = int(future_weeks)
        else:
            future_weeks = int(range_val.replace('w', ''))
            
        # Cap maximum forecast length to avoid unreliable long-term predictions
        future_weeks = min(future_weeks, 52)  # Cap at 1 year
    except ValueError:
        return JsonResponse({'error': 'Invalid range parameter.'}, status=400)

    # Apply filters to queryset
    queryset = EggFarm.objects.all().order_by('production_date')
    
    if farm:
        queryset = queryset.filter(farm_name=farm)
    if site:
        queryset = queryset.filter(site_name=site)
    if breed:
        queryset = queryset.filter(breed=breed)
    if health:
        queryset = queryset.filter(health_issues=health)

    # Group data by week with more comprehensive aggregations
    weekly_data = (
        queryset
        .filter(production_date__isnull=False)  # Ensure we have dates
        .annotate(week=TruncWeek('production_date'))
        .values('week')
        .annotate(
            total=Sum('weekly_egg_production'),
            avg_production=Avg('weekly_egg_production'),
            farm_count=Count('id', distinct=True),
            std_dev=StdDev('weekly_egg_production'),
            min_production=Min('weekly_egg_production'),
            max_production=Max('weekly_egg_production')
        )
        .order_by('week')
    )

    # For longer-term forecasts, require more historical data
    min_required_weeks = min(8, max(4, int(future_weeks / 2)))
    
    if not weekly_data or len(weekly_data) < min_required_weeks:
        return JsonResponse({
            'labels': [], 
            'actual': [], 
            'forecasts': {},
            'error': f'Insufficient historical data for a {future_weeks}-week forecast. Need at least {min_required_weeks} weeks of historical data.'
        })

    # Convert to pandas DataFrame for forecasting
    df = pd.DataFrame(list(weekly_data))
    df.rename(columns={'week': 'ds', 'total': 'y'}, inplace=True)
    df['ds'] = pd.to_datetime(df['ds'])
    df = df.sort_values('ds')
    
    # Log the dataframe for debugging
    logger.info(f"Forecast dataframe before processing: {df.head()}")
    
    # Detect and handle outliers using IQR method
    Q1 = df['y'].quantile(0.25)
    Q3 = df['y'].quantile(0.75)
    IQR = Q3 - Q1
    
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    # Identify outliers
    outliers = df[(df['y'] < lower_bound) | (df['y'] > upper_bound)]
    if not outliers.empty:
        logger.info(f"Detected {len(outliers)} outliers in the dataset")
        
        # Replace outliers with moving average or median for better forecasting
        for idx in outliers.index:
            window = df.index.get_loc(idx)
            start = max(0, window - 2)
            end = min(len(df), window + 3)
            nearby_values = df.iloc[start:end]['y'].values
            nearby_values = nearby_values[nearby_values >= lower_bound]
            nearby_values = nearby_values[nearby_values <= upper_bound]
            
            if len(nearby_values) > 0:
                df.at[idx, 'y'] = np.median(nearby_values)
            else:
                df.at[idx, 'y'] = df['y'].median()
    
    # Fill any missing values with better interpolation
    
    df['ds'] = pd.to_datetime(df['ds'])
    df.set_index('ds', inplace=True)
    if not isinstance(df.index, pd.DatetimeIndex):
        df.index = pd.to_datetime(df.index)
    df['y'] = df['y'].interpolate(method='time')
    df.reset_index(inplace=True)

    
    # Add sequential index for regression
    df['x'] = range(len(df))
    
    # Add time-based features for better forecasting
    df['month'] = df['ds'].dt.month
    df['quarter'] = df['ds'].dt.quarter
    df['day_of_year'] = df['ds'].dt.dayofyear
    
    # Apply min/max normalization to actual values to avoid extreme forecasts
    historical_min = df['y'].min() * 0.85
    historical_max = df['y'].max() * 1.15
    historical_mean = df['y'].mean()
    historical_std = df['y'].std()

    # Common variables for all forecasting methods
    last_date = df['ds'].max()
    future_dates = [last_date + timedelta(weeks=i) for i in range(1, future_weeks + 1)]
    future_x = np.array(range(len(df), len(df) + future_weeks)).reshape(-1, 1)
    
    # Create future dataframe with time features for advanced models
    future_df = pd.DataFrame({
        'ds': future_dates,
        'x': range(len(df), len(df) + future_weeks)
    })
    future_df['month'] = future_df['ds'].dt.month
    future_df['quarter'] = future_df['ds'].dt.quarter
    future_df['day_of_year'] = future_df['ds'].dt.dayofyear
    
    # Store all forecasts and their confidence intervals
    all_forecasts = {}
    
    # 1. Linear Regression Model (with time features)
    try:
        # Use multiple features for linear regression
        X_train = df[['x', 'month', 'quarter']]
        y_train = df['y']
        
        # Use Ridge regression for better stability
        linear_model = Ridge(alpha=1.0)
        linear_model.fit(X_train, y_train)
        
        # Predict using future features
        X_future = future_df[['x', 'month', 'quarter']]
        linear_forecast = linear_model.predict(X_future)
        
        # Calculate residuals for confidence intervals
        linear_predictions = linear_model.predict(X_train)
        linear_residuals = y_train - linear_predictions
        linear_rmse = np.sqrt(np.mean(linear_residuals**2))
        
        # Generate confidence intervals with more appropriate width
        confidence_width = 1.96 * linear_rmse * np.sqrt(1 + 1/len(df))
        linear_upper = linear_forecast + confidence_width
        linear_lower = linear_forecast - confidence_width
        linear_lower = np.maximum(linear_lower, historical_min)
        
        # Apply min/max constraints
        linear_forecast = np.maximum(np.minimum(linear_forecast, historical_max * 1.2), historical_min)
        
        # Calculate error metrics for model evaluation
        linear_mae = mean_absolute_error(y_train, linear_predictions)
        linear_mape = mean_absolute_percentage_error(y_train, linear_predictions) * 100
        
        all_forecasts['linear'] = {
            'name': 'Enhanced Linear Regression',
            'forecast': [None] * len(df) + linear_forecast.round(2).tolist(),
            'upper_bound': [None] * len(df) + linear_upper.round(2).tolist(),
            'lower_bound': [None] * len(df) + linear_lower.round(2).tolist(),
            'stats': {
                'rmse': round(float(linear_rmse), 2),
                'mae': round(float(linear_mae), 2),
                'mape': round(float(linear_mape), 2)
            }
        }
    except Exception as e:
        logger.error(f"Enhanced linear regression error: {str(e)}")
    
    # 2. Polynomial Regression with cross-validation
    try:
        # Find optimal polynomial degree using cross-validation
        degrees = [1, 2, 3]
        best_degree = 1
        best_cv_score = float('inf')
        
        for degree in degrees:
            poly_features = PolynomialFeatures(degree=degree)
            X_poly = poly_features.fit_transform(df[['x']])
            
            # Use cross-validation to find best model
            cv_scores = -cross_val_score(
                LinearRegression(), 
                X_poly, 
                df['y'], 
                cv=min(5, len(df)), 
                scoring='neg_mean_squared_error'
            )
            avg_cv_score = np.mean(cv_scores)
            
            if avg_cv_score < best_cv_score:
                best_cv_score = avg_cv_score
                best_degree = degree
        
        # Create model with best degree
        poly_model = make_pipeline(
            PolynomialFeatures(degree=best_degree),
            Ridge(alpha=1.0)
        )
        poly_model.fit(df[['x']], df['y'])
        
        # Generate predictions
        poly_forecast = poly_model.predict(future_x)
        
        # Calculate confidence intervals
        poly_predictions = poly_model.predict(df[['x']].values)
        poly_residuals = df['y'] - poly_predictions
        poly_rmse = np.sqrt(np.mean(poly_residuals**2))
        
        # Generate confidence intervals with increasing width for future predictions
        confidence_widths = [
            1.96 * poly_rmse * np.sqrt(1 + (i+1)/len(df)) 
            for i in range(future_weeks)
        ]
        
        poly_upper = poly_forecast + confidence_widths
        poly_lower = poly_forecast - confidence_widths
        poly_lower = np.maximum(poly_lower, historical_min)
        
        # Apply constraints to prevent unrealistic growth/decline
        growth_limit = 0.15  # 15% maximum change per week
        for i in range(1, len(poly_forecast)):
            max_change = poly_forecast[i-1] * growth_limit
            poly_forecast[i] = max(
                poly_forecast[i-1] - max_change,
                min(poly_forecast[i], poly_forecast[i-1] + max_change)
            )
        
        # Apply min/max constraints
        poly_forecast = np.maximum(np.minimum(poly_forecast, historical_max * 1.2), historical_min)
        
        # Calculate error metrics
        poly_mae = mean_absolute_error(df['y'], poly_predictions)
        poly_mape = mean_absolute_percentage_error(df['y'], poly_predictions) * 100
        
        all_forecasts['polynomial'] = {
            'name': f'Polynomial Regression (degree={best_degree})',
            'forecast': [None] * len(df) + poly_forecast.round(2).tolist(),
            'upper_bound': [None] * len(df) + poly_upper.round(2).tolist(),
            'lower_bound': [None] * len(df) + poly_lower.round(2).tolist(),
            'stats': {
                'rmse': round(float(poly_rmse), 2),
                'mae': round(float(poly_mae), 2),
                'mape': round(float(poly_mape), 2)
            }
        }
    except Exception as e:
        logger.error(f"Polynomial regression error: {str(e)}")
    
    # 3. Improved Holt-Winters Exponential Smoothing with automatic parameter selection
    if len(df) >= 6:
        try:
            # Test different model configurations
            hw_configs = [
                {'trend': 'add', 'seasonal': None, 'damped': True},
                {'trend': 'add', 'seasonal': None, 'damped': False},
                {'trend': 'mul', 'seasonal': None, 'damped': True}
            ]
            
            best_hw_aic = float('inf')
            best_hw_model = None
            best_config = None
            
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore")
                
                for config in hw_configs:
                    try:
                        hw_test = ExponentialSmoothing(
                            df['y'],
                            **config
                        ).fit(optimized=True)
                        
                        if hasattr(hw_test, 'aic') and hw_test.aic < best_hw_aic:
                            best_hw_aic = hw_test.aic
                            best_hw_model = hw_test
                            best_config = config
                    except:
                        continue
            
            if best_hw_model is None:
                # Fallback to simplest model if optimization fails
                best_hw_model = ExponentialSmoothing(
                    df['y'],
                    trend='add',
                    seasonal=None,
                    damped=True
                ).fit()
            
            # Generate forecast
            hw_forecast = best_hw_model.forecast(future_weeks)
            
            # Calculate prediction intervals
            hw_residuals = df['y'] - best_hw_model.fittedvalues
            hw_rmse = np.sqrt(np.mean(hw_residuals**2))
            
            # Generate confidence intervals with increasing width
            hw_upper = np.array([
                hw_forecast[i] + (hw_rmse * 1.96 * np.sqrt(1 + (i+1)/len(df)))
                for i in range(future_weeks)
            ])
            
            hw_lower = np.array([
                max(hw_forecast[i] - (hw_rmse * 1.96 * np.sqrt(1 + (i+1)/len(df))), historical_min)
                for i in range(future_weeks)
            ])
            
            # Apply min/max constraints with smoothing for more realistic values
            for i in range(len(hw_forecast)):
                hw_forecast[i] = max(min(hw_forecast[i], historical_max * 1.2), historical_min)
                
                # Apply smoothing for longer forecasts
                if i > 0:
                    # Blend with previous value to avoid jumps
                    smoothing_factor = min(0.2 + (i * 0.01), 0.5)  # Increase smoothing over time
                    hw_forecast[i] = (1 - smoothing_factor) * hw_forecast[i] + smoothing_factor * hw_forecast[i-1]
            
            # Calculate error metrics
            hw_mae = mean_absolute_error(df['y'].iloc[1:], best_hw_model.fittedvalues[:-1])
            hw_mape = mean_absolute_percentage_error(df['y'].iloc[1:], best_hw_model.fittedvalues[:-1]) * 100
            
            all_forecasts['holt_winters'] = {
                'name': 'Exponential Smoothing',
                'forecast': [None] * len(df) + hw_forecast.round(2).tolist(),
                'upper_bound': [None] * len(df) + hw_upper.round(2).tolist(),
                'lower_bound': [None] * len(df) + hw_lower.round(2).tolist(),
                'stats': {
                    'rmse': round(float(hw_rmse), 2),
                    'mae': round(float(hw_mae), 2),
                    'mape': round(float(hw_mape), 2),
                    'params': str(best_config)
                }
            }
        except Exception as e:
            logger.error(f"Holt-Winters error: {str(e)}")
    
    # 4. ARIMA model (new addition)
    if len(df) >= 10:  # Need more data for ARIMA
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore")
                
                # Use auto_arima to find best parameters
                arima_model = auto_arima(
                    df['y'],
                    start_p=0, start_q=0,
                    max_p=3, max_q=3, max_d=2,
                    seasonal=False,
                    trace=False,
                    error_action='ignore',
                    suppress_warnings=True,
                    stepwise=True
                )
                
                # Generate forecast
                arima_forecast, conf_int = arima_model.predict(
                    n_periods=future_weeks,
                    return_conf_int=True,
                    alpha=0.05
                )
                
                # Extract upper and lower bounds
                arima_lower = conf_int[:, 0]
                arima_upper = conf_int[:, 1]
                
                # Calculate residuals
                arima_residuals = df['y'] - arima_model.predict_in_sample()
                arima_rmse = np.sqrt(np.mean(arima_residuals**2))
                
                # Apply min/max constraints
                arima_forecast = np.maximum(np.minimum(arima_forecast, historical_max * 1.2), historical_min)
                arima_lower = np.maximum(arima_lower, historical_min)
                arima_upper = np.minimum(arima_upper, historical_max * 1.3)
                
                # Calculate error metrics
                arima_mae = mean_absolute_error(df['y'], arima_model.predict_in_sample())
                arima_mape = mean_absolute_percentage_error(df['y'], arima_model.predict_in_sample()) * 100
                
                all_forecasts['arima'] = {
                    'name': f'ARIMA {arima_model.order}',
                    'forecast': [None] * len(df) + arima_forecast.round(2).tolist(),
                    'upper_bound': [None] * len(df) + arima_upper.round(2).tolist(),
                    'lower_bound': [None] * len(df) + arima_lower.round(2).tolist(),
                    'stats': {
                        'rmse': round(float(arima_rmse), 2),
                        'mae': round(float(arima_mae), 2),
                        'mape': round(float(arima_mape), 2),
                        'order': str(arima_model.order)
                    }
                }
        except Exception as e:
            logger.error(f"ARIMA error: {str(e)}")
    
    # 5. Prophet model (new addition)
    if len(df) >= 8:
        try:
            # Prepare dataframe for Prophet
            prophet_df = df[['ds', 'y']].copy()
            
            # Add changepoint_prior_scale based on data volatility
            volatility = df['y'].std() / df['y'].mean()
            changepoint_prior = 0.05 + (volatility * 0.2)  # Adjust based on volatility
            
            # Create and fit Prophet model
            prophet_model = Prophet(
                changepoint_prior_scale=changepoint_prior,
                yearly_seasonality='auto',
                weekly_seasonality=False,
                daily_seasonality=False
            )
            
            # Add monthly seasonality
            prophet_model.add_seasonality(
                name='monthly',
                period=30.5,
                fourier_order=2
            )
            
            prophet_model.fit(prophet_df)
            
            # Create future dataframe
            prophet_future = prophet_model.make_future_dataframe(
                periods=future_weeks,
                freq='W'
            )
            
            # Generate forecast
            prophet_forecast = prophet_model.predict(prophet_future)
            
            # Extract values for the future period
            prophet_values = prophet_forecast['yhat'].iloc[-future_weeks:].values
            prophet_lower = prophet_forecast['yhat_lower'].iloc[-future_weeks:].values
            prophet_upper = prophet_forecast['yhat_upper'].iloc[-future_weeks:].values
            
            # Apply min/max constraints
            prophet_values = np.maximum(np.minimum(prophet_values, historical_max * 1.2), historical_min)
            prophet_lower = np.maximum(prophet_lower, historical_min)
            prophet_upper = np.minimum(prophet_upper, historical_max * 1.3)
            
            # Calculate error metrics on historical data
            prophet_historical = prophet_forecast['yhat'].iloc[:-future_weeks]
            prophet_rmse = np.sqrt(mean_squared_error(df['y'], prophet_historical))
            prophet_mae = mean_absolute_error(df['y'], prophet_historical)
            prophet_mape = mean_absolute_percentage_error(df['y'], prophet_historical) * 100
            
            all_forecasts['prophet'] = {
                'name': 'Prophet',
                'forecast': [None] * len(df) + prophet_values.round(2).tolist(),
                'upper_bound': [None] * len(df) + prophet_upper.round(2).tolist(),
                'lower_bound': [None] * len(df) + prophet_lower.round(2).tolist(),
                'stats': {
                    'rmse': round(float(prophet_rmse), 2),
                    'mae': round(float(prophet_mae), 2),
                    'mape': round(float(prophet_mape), 2)
                }
            }
        except Exception as e:
            logger.error(f"Prophet error: {str(e)}")
    
    # 6. Create weighted ensemble forecast
    if all_forecasts:
        try:
            # Get all available forecasts
            available_forecasts = {}
            for model_name, model_data in all_forecasts.items():
                if model_name != 'ensemble' and len(model_data['forecast'][len(df):]) == future_weeks:
                    available_forecasts[model_name] = np.array(model_data['forecast'][len(df):])
            
            if available_forecasts:
                # Weight models by the inverse of their error metrics
                # Lower error = higher weight
                weights = {}
                total_weight = 0
                
                for model_name in available_forecasts.keys():
                    # Use inverse of RMSE as weight
                    model_rmse = all_forecasts[model_name]['stats']['rmse']
                    model_weight = 1 / (model_rmse + 0.1)  # Add small constant to avoid division by zero
                    weights[model_name] = model_weight
                    total_weight += model_weight
                
                # Normalize weights
                for model_name in weights:
                    weights[model_name] /= total_weight
                
                # Calculate weighted ensemble forecast
                ensemble_forecast = np.zeros(future_weeks)
                upper_ensemble = np.zeros(future_weeks)
                lower_ensemble = np.zeros(future_weeks)
                
                for model_name, forecast in available_forecasts.items():
                    weight = weights[model_name]
                    ensemble_forecast += forecast * weight
                    
                    if len(all_forecasts[model_name]['upper_bound'][len(df):]) == future_weeks:
                        upper_ensemble += np.array(all_forecasts[model_name]['upper_bound'][len(df):]) * weight
                    
                    if len(all_forecasts[model_name]['lower_bound'][len(df):]) == future_weeks:
                        lower_ensemble += np.array(all_forecasts[model_name]['lower_bound'][len(df):]) * weight
                
                # Apply smoothing to ensemble forecast
                for i in range(1, len(ensemble_forecast)):
                    # Calculate maximum reasonable change
                    max_change_pct = 0.10  # 10% maximum week-over-week change
                    max_change = ensemble_forecast[i-1] * max_change_pct
                    
                    # Limit change to be within reasonable bounds
                    if ensemble_forecast[i] > ensemble_forecast[i-1] + max_change:
                        ensemble_forecast[i] = ensemble_forecast[i-1] + max_change
                    elif ensemble_forecast[i] < ensemble_forecast[i-1] - max_change:
                        ensemble_forecast[i] = ensemble_forecast[i-1] - max_change
                
                # Apply min/max constraints
                ensemble_forecast = np.maximum(np.minimum(ensemble_forecast, historical_max * 1.2), historical_min)
                
                # Calculate weighted average of error metrics
                weighted_rmse = sum(all_forecasts[m]['stats']['rmse'] * weights[m] for m in weights)
                weighted_mae = sum(all_forecasts[m]['stats']['mae'] * weights[m] for m in weights)
                
                # Get best model by RMSE
                best_model = min(all_forecasts.items(), key=lambda x: x[1]['stats'].get('rmse', float('inf')))[0]
                
                all_forecasts['ensemble'] = {
                    'name': 'Weighted Ensemble',
                    'forecast': [None] * len(df) + ensemble_forecast.round(2).tolist(),
                    'upper_bound': [None] * len(df) + upper_ensemble.round(2).tolist(),
                    'lower_bound': [None] * len(df) + lower_ensemble.round(2).tolist(),
                    'stats': {
                        'rmse': round(float(weighted_rmse), 2),
                        'mae': round(float(weighted_mae), 2),
                        'contributors': list(weights.keys()),
                        'weights': {k: round(v, 2) for k, v in weights.items()},
                        'best_model': best_model
                    }
                }
        except Exception as e:
            logger.error(f"Ensemble forecast error: {str(e)}")
    
    # Check if we have any forecasts
    if not all_forecasts:
        return JsonResponse({
            'error': 'Could not generate forecasts with the available data. Try adding more historical data points.'
        })
    
    # Select which method to return based on request parameter
    if method != 'all' and method in all_forecasts:
        selected_forecasts = {method: all_forecasts[method]}
    elif method == 'best':
        # Find model with lowest RMSE
        best_model = min(all_forecasts.items(), key=lambda x: x[1]['stats'].get('rmse', float('inf')))[0]
        selected_forecasts = {best_model: all_forecasts[best_model]}
    else:
        selected_forecasts = all_forecasts
    
    # Create combined timeline for chart
    labels = df['ds'].dt.strftime('%Y-%m-%d').tolist() + [d.strftime('%Y-%m-%d') for d in future_dates]
    
    # Create datasets for chart
    actual_data = df['y'].round(2).tolist()
    
    # Add farm count information for tooltip context
    farm_counts = df['farm_count'].tolist() + [None] * len(future_dates)
    
    # Append metadata for better interpretation
    metadata = {
        'historical_stats': {
            'min': round(float(historical_min), 2),
            'max': round(float(historical_max), 2),
            'mean': round(float(historical_mean), 2),
            'std_dev': round(float(historical_std), 2),
            'weeks_of_data': len(df),
            'projection_weeks': future_weeks
        }
    }

    return JsonResponse({
        'labels': labels,
        'actual': actual_data,
        'forecasts': selected_forecasts,
        'farm_counts': farm_counts,
        'filters': {
            'farm': farm,
            'site': site,
            'breed': breed,
            'health': health
        },
        'metadata': metadata
    })

def prophet_forecasting(df, periods):
    """Prophet forecasting method"""
    from prophet import Prophet
    
    try:
        # Copy the dataframe to avoid modifying the original
        prophet_df = df.copy()
        
        # Prophet requires 'ds' and 'y' columns
        model = Prophet(
            weekly_seasonality=True,
            daily_seasonality=False,
            yearly_seasonality=True,
            seasonality_mode='multiplicative',
            interval_width=0.95
        )
        
        # Add seasonality if enough data
        if len(prophet_df) >= 52:  # At least a year of data
            model.add_seasonality(name='yearly', period=365.25, fourier_order=5)
        
        # Fit the model
        model.fit(prophet_df)
        
        # Make future dataframe and predict
        future = model.make_future_dataframe(periods=periods, freq='W-MON')
        forecast = model.predict(future)
        
        # Calculate RMSE on training data
        y_true = prophet_df['y']
        y_pred = model.predict(prophet_df[['ds']])['yhat']
        rmse = round(((y_true - y_pred) ** 2).mean() ** 0.5, 2)
        
        # Return results
        result = {
            'name': 'Prophet',
            'forecast': [round(float(y), 2) for y in forecast['yhat'][-periods:].tolist()],
            'upper_bound': [round(float(y), 2) for y in forecast['yhat_upper'][-periods:].tolist()],
            'lower_bound': [round(float(y), 2) for y in forecast['yhat_lower'][-periods:].tolist()],
            'stats': {'rmse': rmse}
        }
        return result
    except Exception as e:
        return {
            'name': 'Prophet (Failed)',
            'forecast': [None] * periods,
            'upper_bound': [None] * periods,
            'lower_bound': [None] * periods,
            'stats': {'error': str(e)}
        }


def arima_forecasting(df, periods):
    """ARIMA forecasting method"""
    try:
        from statsmodels.tsa.arima.model import ARIMA
        import numpy as np
        
        # Copy the dataframe to avoid modifying the original
        arima_df = df.copy()
        
        # Extract the y values for ARIMA
        y = arima_df['y'].values
        
        # Find optimal parameters
        p, d, q = 5, 1, 0  # Default parameters
        
        # Try to find optimal parameters if enough data
        if len(y) > 30:
            from pmdarima import auto_arima
            try:
                auto_model = auto_arima(
                    y, 
                    start_p=1, start_q=1,
                    max_p=5, max_q=5,
                    d=1, max_d=2,
                    seasonal=True,
                    m=52,  # Weekly data
                    information_criterion='aic',
                    trace=False,
                    error_action='ignore',
                    suppress_warnings=True,
                    stepwise=True
                )
                p, d, q = auto_model.order
            except:
                # Fall back to default parameters
                pass
        
        # Fit ARIMA model
        model = ARIMA(y, order=(p, d, q))
        model_fit = model.fit()
        
        # Make predictions
        predictions = model_fit.forecast(steps=periods)
        
        # Calculate confidence intervals (95%)
        from scipy import stats
        stderr = np.sqrt(model_fit.params.var())
        conf_int = stats.norm.ppf(0.975) * stderr
        
        # Create upper and lower bounds
        upper_bound = [max(0, pred + conf_int) for pred in predictions]
        lower_bound = [max(0, pred - conf_int) for pred in predictions]
        
        # Calculate RMSE on training data
        y_pred = model_fit.predict(start=1, end=len(y))
        rmse = round(np.sqrt(np.mean((y[1:] - y_pred) ** 2)), 2)
        
        # Return results
        result = {
            'name': 'ARIMA',
            'forecast': [round(float(pred), 2) for pred in predictions],
            'upper_bound': [round(float(ub), 2) for ub in upper_bound],
            'lower_bound': [round(float(lb), 2) for lb in lower_bound],
            'stats': {'rmse': rmse}
        }
        return result
    except Exception as e:
        return {
            'name': 'ARIMA (Failed)',
            'forecast': [None] * periods,
            'upper_bound': [None] * periods,
            'lower_bound': [None] * periods,
            'stats': {'error': str(e)}
        }


def holt_winters_forecasting(df, periods):
    """Holt-Winters Exponential Smoothing forecasting method"""
    try:
        from statsmodels.tsa.holtwinters import ExponentialSmoothing
        import numpy as np
        
        # Copy the dataframe to avoid modifying the original
        hw_df = df.copy()
        
        # Extract the y values for Holt-Winters
        y = hw_df['y'].values
        
        # Set seasonal period
        seasonal_periods = min(52, len(y) // 2)  # Weekly data, but limited by data length
        
        # Fit Holt-Winters model
        model = ExponentialSmoothing(
            y,
            trend='add',
            seasonal='add',
            seasonal_periods=seasonal_periods,
            damped=True
        )
        model_fit = model.fit(optimized=True)
        
        # Make predictions
        predictions = model_fit.forecast(periods)
        
        # Calculate confidence intervals (95%)
        # For Holt-Winters, we use a simpler approach for confidence intervals
        resid = model_fit.resid
        stderr = np.std(resid)
        conf_int = 1.96 * stderr  # 95% confidence
        
        # Create upper and lower bounds
        upper_bound = [max(0, pred + conf_int) for pred in predictions]
        lower_bound = [max(0, pred - conf_int) for pred in predictions]
        
        # Calculate RMSE on training data
        y_pred = model_fit.fittedvalues
        rmse = round(np.sqrt(np.mean((y - y_pred) ** 2)), 2)
        
        # Return results
        result = {
            'name': 'Holt-Winters',
            'forecast': [round(float(pred), 2) for pred in predictions],
            'upper_bound': [round(float(ub), 2) for ub in upper_bound],
            'lower_bound': [round(float(lb), 2) for lb in lower_bound],
            'stats': {'rmse': rmse}
        }
        return result
    except Exception as e:
        return {
            'name': 'Holt-Winters (Failed)',
            'forecast': [None] * periods,
            'upper_bound': [None] * periods,
            'lower_bound': [None] * periods,
            'stats': {'error': str(e)}
        }


def create_ensemble_forecast(forecasts):
    """Create an ensemble forecast by averaging all methods"""
    try:
        # Get valid forecasts (ignore failed ones)
        valid_forecasts = [f for f in forecasts if None not in f['forecast']]
        
        if not valid_forecasts:
            raise ValueError("No valid forecasts available for ensemble")
        
        # Get the number of periods
        periods = len(valid_forecasts[0]['forecast'])
        
        # Calculate average forecast
        ensemble_forecast = []
        ensemble_upper = []
        ensemble_lower = []
        
        for i in range(periods):
            # Get forecasts for this period
            period_forecasts = [f['forecast'][i] for f in valid_forecasts]
            period_upper = [f['upper_bound'][i] for f in valid_forecasts]
            period_lower = [f['lower_bound'][i] for f in valid_forecasts]
            
            # Calculate average
            ensemble_forecast.append(round(sum(period_forecasts) / len(period_forecasts), 2))
            ensemble_upper.append(round(sum(period_upper) / len(period_upper), 2))
            ensemble_lower.append(round(sum(period_lower) / len(period_lower), 2))
        
        # Calculate average RMSE
        rmse_values = [f['stats'].get('rmse', 0) for f in valid_forecasts if 'rmse' in f['stats']]
        avg_rmse = round(sum(rmse_values) / len(rmse_values), 2) if rmse_values else 0
        
        # Return ensemble result
        result = {
            'name': 'Ensemble',
            'forecast': ensemble_forecast,
            'upper_bound': ensemble_upper,
            'lower_bound': ensemble_lower,
            'stats': {'rmse': avg_rmse}
        }
        return result
    except Exception as e:
        periods = len(forecasts[0]['forecast']) if forecasts else 0
        return {
            'name': 'Ensemble (Failed)',
            'forecast': [None] * periods,
            'upper_bound': [None] * periods,
            'lower_bound': [None] * periods,
            'stats': {'error': str(e)}
        }



@critical_view_decorator
@login_required(login_url='login')
def farm_autocomplete(request):
    """
    Provide autocomplete suggestions for farm names.
    """
    term = request.GET.get('term', '')
    farms = EggFarm.objects.filter(
        farm_name__icontains=term
    ).values_list('farm_name', flat=True).distinct()
    
    suggestions = [{'id': farm, 'text': farm} for farm in farms if farm]

    return JsonResponse(suggestions, safe=False)
    """
    Debug view to check forecast data quality
    """
    farms = EggFarm.objects.all()
    total_farms = farms.count()
    farms_with_production = farms.filter(weekly_egg_production__isnull=False, weekly_egg_production__gt=0)
    farms_with_dates = farms.filter(production_date__isnull=False)

    if farms.exists():
        earliest_record = farms.earliest('production_date').production_date if farms_with_dates.exists() else None
        latest_record = farms.latest('production_date').production_date if farms_with_dates.exists() else None
    else:
        earliest_record = latest_record = None

    weekly_data = (
        farms
        .filter(production_date__isnull=False)
        .annotate(week=TruncWeek('production_date'))
        .values('week')
        .annotate(
            total_production=Sum('weekly_egg_production'),
            farm_count=Count('id')
        )
        .order_by('week')
    )

    context = {
        'total_farms': total_farms,
        'farms_with_production': farms_with_production.count(),
        'farms_with_dates': farms_with_dates.count(),
        'earliest_record': earliest_record,
        'latest_record': latest_record,
        'weekly_data': list(weekly_data)
    }

    return JsonResponse(context)